'''
    Welcome to tc_sorter v1.8.0
    Created by Jeter Lin
'''

from openpyxl import load_workbook
from openpyxl import Workbook, formatting, styles
import json
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from src.auto_case_list_gen import auto_case_list_gen
from src.div import *
from src.detail import *
from result_log import Result_log

# the index of the precondition
pre_index = 5


def json_directory(json_name):
    with open('json\\' + json_name) as f:
        return json.load(f)



class Auto_team_sorter:
    def __init__(self, test_case_list, last_week, continue_from=False):
        # print('Initiallizing...')
        self.test_case_list = str(test_case_list)
        self.output_name = test_case_list[:-10] + '_sorted.xlsx'
        self.auto_output_name = test_case_list[:-10] + '_auto.xlsx'
        self.sheet = (load_workbook(self.test_case_list)).active
        # print('{} loaded successfully'.format(self.test_case_list))

        # Loading JSON file
        self.data_sheet = json_directory('sheet_related.json')
        self.keywords = json_directory('keywords.json')
        self.auto_case_list = json_directory('auto_case_id.json')
        self.locked_tcid = json_directory('locked_tcid.json')
        self.auto_packages = json_directory('pkg_case.json')

        # Determine the titles
        if 'STR' in test_case_list.split('_'):
            titles = self.data_sheet['str_titles']
        else:
            titles = self.data_sheet['titles']

        # Loading the resut from last week
        self.last_week_result = (load_workbook(last_week))
        # print('{} loaded successfully'.format(last_week))

        if continue_from == False:
            self.wb = Workbook()
            # self.wb.active
            for name in self.data_sheet['sheet_names']:
                self.wb.create_sheet(
                    name, int((self.data_sheet['sheet_names']).index(name)))
                self.wb[name].append(titles)
            for fail_name in self.data_sheet['fail_case_sheet']:
                self.wb.create_sheet(fail_name, -1)
                self.wb[fail_name].append(titles)
            # print('Output file initiallized')

            # creating workbook for automated cases
            self.auto_wb = Workbook()
            for name in self.data_sheet['auto_sheetname']:
                self.auto_wb.create_sheet(
                    name, int((self.data_sheet['auto_sheetname']).index(name)))
                self.auto_wb[name].append(self.data_sheet['titles'])

        else:
            self.wb = load_workbook(self.output_name)

    def cell_data(self, row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

    def formatter(self, cell_data):
        for _ in range(4):
            cell_data.insert(1, '')

    def last_week_result_dict(self):
        last_week_dict = {}
        last_week = self.last_week_result
        for sheet in last_week.worksheets:
            if sheet.title not in ['Fail Cases China', 'Cases need update', 'Summary']:
                for last_week_row in sheet.iter_rows(max_col=5, values_only=True):
                    last_week_cell = self.cell_data(last_week_row)
                    if last_week_cell[0] != 'Original GM TC ID':
                        last_week_dict[last_week_cell[0]] = last_week_cell[1:]
        return last_week_dict

    def generate_auto_list(self):
        auto_case_list_gen(self.output_name)

    def sorted_manually(self, cell_data, name_and_num):
        # Sorting the case based on organized sheet dict
        locked_tcid = json_directory('locked_tcid.json')
        for sheet_name in locked_tcid:
            if cell_data[0] in locked_tcid[sheet_name]:
                sheet_name = sheet_name.lower()
                self.wb[sheet_name.lower()].append(cell_data)
                name_and_num[sheet_name] += 1
                return True
        return False

    def sorting(self):
        # print('Opening a new sheet...')
        sheet = self.sheet

        k = 1

        auto_name_and_num = {name: 0 for name in self.data_sheet['auto_sheetname']}

        # Iterate through the unprocessd test cases
        # Only getting the first 6 values of each row (tc, precondition, test_steps, expected_result, test_objective, MEC}
        for row in sheet.iter_rows(max_col=6, values_only=True):
            print('Iterate case no. {}'.format(k), end='\r')
            # turn the data into a list
            cell_data = []
            for cell in row:
                if cell is not None:
                    cell_data.append(cell)
                else:
                    cell_data.append('none')

            # adding 'pass/fail', 'Tester', 'Automation_comment', 'bug ID', 'Note' to the list
            self.formatter(cell_data)
            # determine the phone type
            phone_type(cell_data)
            # determine the user type
            user(cell_data)
            # determine online/offline
            connection(cell_data)
            # determine sign-in/sign-out
            sign_status(cell_data)
            k += 1

            # For Automation
            # ==========================================================================================
            if cell_data[0].lower() in [i.lower() for i in self.auto_packages['trailer_suv']]:
                self.auto_wb['trailer_suv'].append(cell_data)
                auto_name_and_num['trailer_suv'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['waa']]:
                self.auto_wb['waa'].append(cell_data)
                auto_name_and_num['waa'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['bt']]:
                self.auto_wb['bt'].append(cell_data)
                auto_name_and_num['bt'] += 1
            
            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['phone_projection_1']]:
                self.auto_wb['phone_projection_1'].append(cell_data)
                auto_name_and_num['phone_projection_1'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['gas_user_1']]:
                self.auto_wb['gas_user_1'].append(cell_data)
                auto_name_and_num['gas_user_1'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['online_in_no_phone_ac']]:
                self.auto_wb['online_in_no_phone_ac'].append(cell_data)
                auto_name_and_num['online_in_no_phone_ac'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['carplay']]:
                self.auto_wb['carplay'].append(cell_data)
                auto_name_and_num['carplay'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['iphone']]:
                self.auto_wb['iphone'].append(cell_data)
                auto_name_and_num['iphone'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['user_online_out']]:
                self.auto_wb['user_online_out'].append(cell_data)
                auto_name_and_num['user_online_out'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['user_offline_in']]:
                self.auto_wb['user_offline_in'].append(cell_data)
                auto_name_and_num['user_offline_in'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['user_offline_out']]:
                self.auto_wb['user_offline_out'].append(cell_data)
                auto_name_and_num['user_offline_out'] += 1      

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['online_in_ac']]:
                self.auto_wb['online_in_ac'].append(cell_data)
                auto_name_and_num['online_in_ac'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['online_out_ac']]:
                self.auto_wb['online_out_ac'].append(cell_data)
                auto_name_and_num['online_out_ac'] += 1

            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['offline_in_ac']]:
                self.auto_wb['offline_in_ac'].append(cell_data)
                auto_name_and_num['offline_in_ac'] += 1
            
            elif cell_data[0].lower() in [i.lower() for i in self.auto_packages['offline_out_ac']]:
                self.auto_wb['offline_out_ac'].append(cell_data)
                auto_name_and_num['offline_out_ac'] += 1

            # ==========================================================================================

        self.auto_wb.save(self.auto_output_name)

        print('===============================================')
        print('[SUMMARY] Auto')
        Result_log(auto_name_and_num).log()
        print('===============================================')


if __name__ == '__main__':
    # __init__(self, test_case_list, last_week, continue_from=False)
    
    # testing = Auto_sorter('W42_STR_cases.xlsx',
    #                     'W41_STR_Sorted.xlsx', continue_from=False)
    # testing.sorting()

    testing1 = Auto_team_sorter('new_one_month_cases.xlsx',
                        'W41_Main_Sorted.xlsx', continue_from=False)
    testing1.sorting()