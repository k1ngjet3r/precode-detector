from openpyxl import load_workbook
import json

wb = load_workbook('new_pkg_design.xlsx')

sheetnames = wb.sheetnames

pkg_design = {}

for name in sheetnames:
    if name not in ['Summary', ' back_to_manual', 'TCID', 'gas_user', 'Taipei AI Scope', 'MY23_scope', 'case_details']:
        tmp = []
        for case in wb[name].iter_rows(max_col=1, values_only=True):
            if case[0] and case[0].lower()[:3] == 'tc_':
                tmp.append(case[0].lower())
        pkg_design[name] = tmp

# print(pkg_design)
sum = 0

for i in pkg_design:
    print('{}: {}'.format(i, len(pkg_design[i])))
    sum += len(pkg_design[i])

print('Summary: {}'.format(sum))

with open('json/pkg_design.json', 'w') as j:
    json.dump(pkg_design, j)